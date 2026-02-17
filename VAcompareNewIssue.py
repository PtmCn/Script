import pandas as pd
import glob
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
import os

def standardize(val):
    if val is None or pd.isna(val):
        return ""
    return str(val).strip().upper()

def apply_header_style(cell):
    """Applies Blue background, White Bold text, and Center alignment."""
    cell.fill = PatternFill(start_color='0070C0', end_color='0070C0', fill_type='solid')
    cell.font = Font(color='FFFFFF', bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center')

def safe_write(ws, row, col, value, style_func=None):
    """Safely write to cells, handling merged ranges if they exist."""
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, openpyxl.cell.cell.MergedCell):
        for merged_range in ws.merged_cells.ranges:
            if cell.coordinate in merged_range:
                cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                break
    cell.value = value
    if style_func: 
        style_func(cell)

def process_new_issues_with_risk():
    output_folder = "output"
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)

    q3_files = glob.glob('*Q3*.xlsx') + glob.glob('*q3*.xlsx')
    q4_files = glob.glob('*Q4*.xlsx') + glob.glob('*q4*.xlsx')

    if not q3_files or not q4_files:
        print("Error: Could not find Q3 and Q4 files.")
        return

    q3_path, q4_path = q3_files[0], q4_files[0]
    # Deduplicate and match based on these 4 columns
    match_cols = ['Plugin ID', 'Host', 'Protocol', 'Port']

    # --- STEP 1: LOAD Q3 INTO LOOKUP SET ---
    q3_dict = pd.read_excel(q3_path, sheet_name=None)
    df_q3_all = pd.concat(q3_dict.values(), ignore_index=True)
    
    q3_existing_issues = set()
    for _, row in df_q3_all.iterrows():
        fingerprint = tuple(standardize(row.get(c)) for c in match_cols)
        if fingerprint[0]:
            q3_existing_issues.add(fingerprint)

    # --- STEP 2: PROCESS Q4 BASE WORKBOOK ---
    wb = openpyxl.load_workbook(q4_path)
    new_issue_fill = PatternFill(start_color='B02418', end_color='B02418', fill_type='solid')
    existing_fill = PatternFill(start_color='4FAD5B', end_color='4FAD5B', fill_type='solid')

    summary_data = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers = [str(cell.value).strip() if cell.value else f"BlankCol_{i}" for i, cell in enumerate(ws[1])]
        
        try:
            # Explicitly mapping the Risk column from your input
            col_map = {col: headers.index(col) + 1 for col in match_cols + ['Name', 'Risk']}
        except ValueError as e:
            continue

        seen_in_q4 = set()
        rows_to_delete = []

        # Deduplicate Q4 internally
        for row_idx in range(2, ws.max_row + 1):
            f_q4 = tuple(standardize(ws.cell(row=row_idx, column=col_map[c]).value) for c in match_cols)
            if f_q4 in seen_in_q4 or ws.cell(row=row_idx, column=col_map['Plugin ID']).value is None:
                rows_to_delete.append(row_idx)
            else:
                seen_in_q4.add(f_q4)

        for row_idx in reversed(rows_to_delete):
            ws.delete_rows(row_idx)

        # Status Comparison
        status_col_idx = col_map['Name'] + 1
        ws.insert_cols(status_col_idx)
        ws.cell(row=1, column=status_col_idx).value = "Status"

        # Dictionary to hold Risk counts for New vs Existing
        counts = {
            'new': {'TOTAL': 0, 'CRITICAL': 0, 'HIGH': 0, 'MEDIUM': 0, 'LOW': 0},
            'existing': {'TOTAL': 0, 'CRITICAL': 0, 'HIGH': 0, 'MEDIUM': 0, 'LOW': 0}
        }

        for row_idx in range(2, ws.max_row + 1):
            check_key = tuple(standardize(ws.cell(row=row_idx, column=col_map[c]).value) for c in match_cols)
            risk_val = standardize(ws.cell(row=row_idx, column=col_map['Risk']).value)
            status_cell = ws.cell(row=row_idx, column=status_col_idx)
            
            # Logic: If in Q4 but NOT in Q3 -> New Issue
            category = 'new' if check_key not in q3_existing_issues else 'existing'
            status_cell.value = "New Issue" if category == 'new' else "Existing"
            status_cell.fill = new_issue_fill if category == 'new' else existing_fill
            
            counts[category]['TOTAL'] += 1
            if risk_val in counts[category]:
                counts[category][risk_val] += 1
        
        summary_data[sheet_name] = counts

    # --- STEP 3: CREATE RecurrenceSummary TAB ---
    if "RecurrenceSummary" in wb.sheetnames:
        del wb["RecurrenceSummary"]
    ws_sum = wb.create_sheet("RecurrenceSummary", 0)

    risks = ['CRITICAL', 'HIGH', 'MEDIUM', 'LOW']
    g_total = {'new': 0, 'ext': 0, 'new_r': [0]*4, 'ext_r': [0]*4}

    # --- TABLE 1: NEW VS EXISTING SCOPE ---
    row_ptr = 2
    safe_write(ws_sum, row_ptr, 2, "New Vulnerability Issue Summary", apply_header_style)
    row_ptr += 1
    safe_write(ws_sum, row_ptr, 2, "Domain", apply_header_style)
    safe_write(ws_sum, row_ptr, 3, "New Issue", apply_header_style)
    safe_write(ws_sum, row_ptr, 4, "Existing Issue", apply_header_style)
    safe_write(ws_sum, row_ptr, 5, "Total Issue", apply_header_style)
    
    row_ptr += 1
    for domain, data in summary_data.items():
        safe_write(ws_sum, row_ptr, 2, domain)
        safe_write(ws_sum, row_ptr, 3, data['new']['TOTAL'] or "-")
        safe_write(ws_sum, row_ptr, 4, data['existing']['TOTAL'] or "-")
        safe_write(ws_sum, row_ptr, 5, (data['new']['TOTAL'] + data['existing']['TOTAL']) or "-")
        
        g_total['new'] += data['new']['TOTAL']
        g_total['ext'] += data['existing']['TOTAL']
        for i, r in enumerate(risks):
            g_total['new_r'][i] += data['new'][r]
            g_total['ext_r'][i] += data['existing'][r]
        row_ptr += 1

    # Total Table 1
    safe_write(ws_sum, row_ptr, 2, "Total", apply_header_style)
    for col, val in enumerate([g_total['new'], g_total['ext'], (g_total['new'] + g_total['ext'])], 3):
        safe_write(ws_sum, row_ptr, col, val or "-", apply_header_style)

    # --- TABLE 2: NEW ISSUES RISK MAPPING (3 row gap) ---
    row_ptr += 4
    safe_write(ws_sum, row_ptr, 2, "[temp] new issues risk summary", apply_header_style)
    row_ptr += 1
    safe_write(ws_sum, row_ptr, 2, "Domain", apply_header_style)
    for i, r in enumerate(risks + ['Total']):
        safe_write(ws_sum, row_ptr, 3+i, r, apply_header_style)
    
    row_ptr += 1
    for domain, data in summary_data.items():
        safe_write(ws_sum, row_ptr, 2, domain)
        for i, r in enumerate(risks):
            safe_write(ws_sum, row_ptr, 3+i, data['new'][r] or "-")
        safe_write(ws_sum, row_ptr, 7, data['new']['TOTAL'] or "-")
        row_ptr += 1
    
    # Total Table 2
    safe_write(ws_sum, row_ptr, 2, "Total", apply_header_style)
    for i in range(4): safe_write(ws_sum, row_ptr, 3+i, g_total['new_r'][i] or "-", apply_header_style)
    safe_write(ws_sum, row_ptr, 7, g_total['new'] or "-", apply_header_style)

    # --- TABLE 3: EXISTING ISSUES RISK MAPPING (3 row gap) ---
    row_ptr += 4
    safe_write(ws_sum, row_ptr, 2, "[temp] existing issues risk summary", apply_header_style)
    row_ptr += 1
    safe_write(ws_sum, row_ptr, 2, "Domain", apply_header_style)
    for i, r in enumerate(risks + ['Total']):
        safe_write(ws_sum, row_ptr, 3+i, r, apply_header_style)
    
    row_ptr += 1
    for domain, data in summary_data.items():
        safe_write(ws_sum, row_ptr, 2, domain)
        for i, r in enumerate(risks):
            safe_write(ws_sum, row_ptr, 3+i, data['existing'][r] or "-")
        safe_write(ws_sum, row_ptr, 7, data['existing']['TOTAL'] or "-")
        row_ptr += 1

    # Total Table 3
    safe_write(ws_sum, row_ptr, 2, "Total", apply_header_style)
    for i in range(4): safe_write(ws_sum, row_ptr, 3+i, g_total['ext_r'][i] or "-", apply_header_style)
    safe_write(ws_sum, row_ptr, 7, g_total['ext'] or "-", apply_header_style)

    # Formatting
    ws_sum.column_dimensions['B'].width = 35
    for col in ['C', 'D', 'E', 'F', 'G']:
        ws_sum.column_dimensions[col].width = 15

    wb.save(os.path.join(output_folder, "VA_New_Issues_Final.xlsx"))
    print("\n--- Success! New Issue Summary with Risk Mapping Created. ---")

if __name__ == "__main__":
    process_new_issues_with_risk()